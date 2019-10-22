import json
import psycopg2
import datetime
from simple_salesforce import Salesforce
from openpyxl import Workbook

# JSON config variables
with open('config.json') as json_config:
    config = json.load(json_config)

# Database setup
connection = psycopg2.connect(
    user=config["database"]["user"], password=config["database"]["password"], database=config["database"]["database"])
cursor = connection.cursor()

# Salesforce setup
sf = Salesforce(username=config["salesforce"]["username"],
                password=config["salesforce"]["password"], security_token=config["salesforce"]["token"])
partner_opportunity_query = "SELECT Partner_Owner__c, convertCurrency(Amount) FROM Opportunity WHERE IsClosed = False and Partner_Owner__c != null"
records = sf.query(partner_opportunity_query)['records']

# Create Excel workbook
wb = Workbook()
count_ws = wb.active
count_ws.title = "Count"
sum_ws = wb.create_sheet("Sum")

#To store partner ids
partner_dictionary = {}

for sheet in wb:
    sheet["A1"] = "Date"

file_name = "/Users/alec/Desktop/partnerhistoricalreport.xlsx"
wb.save(file_name)

# Select all unique dates in ascending order
cursor.execute("SELECT DISTINCT(record_date) FROM partners ORDER BY record_date ASC")
record_dates = cursor.fetchall()

# For each date, select records
for selected_date in record_dates:
    
    current_sheet_row = record_dates.index(selected_date) + 2 
    
    for sheet in wb:
        sheet["A" + str(current_sheet_row)] = str(selected_date[0])


    cursor.execute("SELECT * FROM partners WHERE record_date = '" + str(selected_date[0]) + "'")
    
    current_records = cursor.fetchall()

    for record in current_records:

        partner_id = record[0]
        partner_count = record[1]
        partner_sum = record[2]

        if partner_id not in partner_dictionary:
            partner_dictionary[partner_id] = len(partner_dictionary) + 2


            partner_column = partner_dictionary[partner_id]

            for sheet in wb:
                sheet.cell(row=1, column=partner_column, value=partner_id)
            
            count_ws.cell(row=current_sheet_row, column=partner_column, value=partner_count)
            sum_ws.cell(row=current_sheet_row, column=partner_column, value=partner_sum)


        else:
            partner_column = partner_dictionary[partner_id]
            count_ws.cell(row=current_sheet_row, column=partner_column, value=partner_count)
            sum_ws.cell(row=current_sheet_row, column=partner_column, value=partner_sum)

wb.save(file_name)

#Replace IDs with account names from Salesforce 

for account_id in partner_dictionary:

    account_name = sf.query("SELECT Name FROM Account WHERE Id = '" + account_id +"'")["records"][0]["Name"]
    partner_column = partner_dictionary[account_id]

    for sheet in wb:
        sheet.cell(row=1, column=partner_column, value=account_name)


wb.save(file_name)